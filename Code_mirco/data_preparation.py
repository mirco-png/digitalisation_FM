import os #data structure
import csv
import pandas as pd
import openpyxl

folder_path = os.path.join(os.path.dirname(os.path.abspath(__file__))+"\data", "")
wb = openpyxl.Workbook() 
sheet = wb.active 
sheet.cell(row = 1, column = 1).value = ' hello '
sheet.cell(row = 2, column = 2).value = ' everyone '
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save(f"{folder_path}test.xlsx")

#folder_path = "./data/own_data/Midfield"



